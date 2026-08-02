import os
import shutil
from typing import Dict, Any, List
from copy import copy

import polars as pl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from PySide6.QtCore import QObject, Signal


class Writer(QObject):
    """
    Writes updated data to a new file with minimal memory usage.
    
    Supports XLSX (via xlsxwriter for speed) and CSV output.
    Preserves header formatting and column widths (XLSX only).
    Restores formulas after data update (XLSX only).
    """
    
    progress_updated = Signal(str)
    
    def __init__(
        self, session_data: Dict[str, Any], diff_result: Dict[str, Any]
    ):
        """Initialize the writer with session and diff data."""
        super().__init__()
        self.source_path = session_data['source_path']
        self.target_path = session_data['target_path']
        self.output_path = session_data['output_path']
        self.column_mapping = session_data['column_mapping']
        self.target_sku_col = session_data['target_sku_col']
        self.formula_info = session_data.get('formula_info', {})
        
        self.updates = diff_result.get('updates', [])
        self.additions = diff_result.get('additions')
        self.new_columns_data = diff_result.get('new_columns_data', {})
        
        self._build_update_lookup()
        self._build_new_columns_lookup()
    
    def _build_update_lookup(self):
        """Build lookup dictionary for efficient row updates."""
        self.updates_by_sku = {}
        for update in self.updates:
            sku = update['sku']
            if sku not in self.updates_by_sku:
                self.updates_by_sku[sku] = {}
            self.updates_by_sku[sku][update['column']] = update['new_value']
        
        self.update_skus = set(self.updates_by_sku.keys())
        self.update_columns = list(
            self.column_mapping.get('updates', {}).values()
        )
    
    def _build_new_columns_lookup(self):
        """Build lookup dictionary for new column values by row index."""
        new_columns_mapping = self.column_mapping.get('new_columns', {})
        self.new_columns_combined = None  # Initialize to avoid attribute error
        if not new_columns_mapping or not self.new_columns_data:
            return
        
        # Combine all new column DataFrames into one wide DataFrame
        dfs = list(self.new_columns_data.values())
        combined = dfs[0]
        for df in dfs[1:]:
            combined = combined.join(df, on='row_idx', how='left')
        self.new_columns_combined = combined
    
    def write(self):
        """Execute the complete writing process."""
        is_csv = self.output_path.endswith('.csv')
        
        if is_csv:
            temp_path = self.output_path.replace('.csv', '_temp.csv')
        else:
            temp_path = self.output_path.replace('.xlsx', '_temp.xlsx')
        
        try:
            if is_csv:
                self._write_csv(temp_path)
            else:
                self._write_xlsx(temp_path)
                if self.formula_info.get('has_formulas'):
                    self._restore_formulas(temp_path)
            
            if os.path.exists(self.output_path):
                os.remove(self.output_path)
            shutil.move(temp_path, self.output_path)
        
        except Exception as e:
            if os.path.exists(temp_path):
                os.remove(temp_path)
            raise e
        
        finally:
            self.updates_by_sku.clear()
    
    # ── Optimized: single join via Polars pivot (eager pivot for stability) ──
    def _apply_updates(self, source_lf):
        """Apply all column updates in a single join using Polars pivot.
        
        Converts the flat updates list into a wide DataFrame via pivot,
        then joins once to the source. Uses eager pivot for maximum
        compatibility across Polars versions.
        
        Args:
            source_lf: Source Polars LazyFrame.
            
        Returns:
            LazyFrame with all updates applied.
        """
        
        if not self.updates:
            return source_lf
        
        sku_col = self.target_sku_col
        
        # Build update DataFrame directly from updates list (eager)
        updates_df = pl.DataFrame(self.updates)
        
        # Eager pivot (columns parameter is stable across all versions)
        updates_wide = (
            updates_df.pivot(
                index='sku',
                columns='column',
                values='new_value'
            )
            .rename({'sku': sku_col})
            .lazy()
        )
        
        # Add prefix to update columns to avoid name collision with source
        # Use collect_schema().names() to avoid PerformanceWarning
        update_cols = [c for c in updates_wide.collect_schema().names() if c != sku_col]
        rename_map = {c: f"__upd_{c}" for c in update_cols}
        updates_wide = updates_wide.rename(rename_map)
        
        # Single left join
        source_lf = source_lf.join(updates_wide, on=sku_col, how='left')
        
        # Coalesce: use updated value if not null, otherwise keep original
        coalesce_exprs = []
        # Resolve schema once for efficient column name checks
        source_schema = source_lf.collect_schema()
        source_col_names = source_schema.names()
        for col in update_cols:
            upd_col = f"__upd_{col}"
            if upd_col in source_col_names:
                coalesce_exprs.append(
                    pl.when(pl.col(upd_col).is_not_null())
                    .then(pl.col(upd_col))
                    .otherwise(pl.col(col))
                    .alias(col)
                )
        
        if coalesce_exprs:
            source_lf = source_lf.with_columns(coalesce_exprs)
            # schema changed, we'll re-fetch when needed
        
        # Drop temporary update columns
        # Use collect_schema().names() to get current column names without warning
        temp_cols = [c for c in source_lf.collect_schema().names() if c.startswith("__upd_")]
        if temp_cols:
            source_lf = source_lf.drop(temp_cols)
        
        return source_lf
    
    # ── XLSX Writing ────────────────────────────────────────────────────
    def _get_header_formats(self):
        """Extract header formatting from the original target file.
        Uses read-only mode for speed. Column widths are intentionally
        skipped due to openpyxl read-only limitations.
        """
        wb = load_workbook(self.target_path, read_only=True)
        ws = wb.active

        header_formats = {}
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            if cell.has_style:
                font = cell.font
                fill = cell.fill
                align = cell.alignment
                header_formats[col_idx] = {
                    'font_name': font.name,
                    'font_size': font.size,
                    'bold': font.bold,
                    'font_color': font.color.rgb if font.color else None,
                    'bg_color': fill.start_color.rgb if fill and fill.start_color else None,
                    'align': align.horizontal,
                    'valign': align.vertical,
                    'num_format': cell.number_format,
                }

        wb.close()
        return header_formats

    def _write_xlsx(self, output_path: str):
        """Write data to XLSX using Polars + xlsxwriter with low memory footprint.
        
        Writes to a temporary Parquet file first (streaming) to reduce memory,
        then streams from Parquet to xlsxwriter in chunks. Header formatting
        is applied directly during write for maximum speed.
        """
        self.progress_updated.emit("Building output data...")
        
        # Read target file as LazyFrame (if possible)
        source_lf = self._get_lazy_source()
        source_lf = self._apply_updates(source_lf)
        
        # Add new columns (optimized: single join)
        new_columns_mapping = self.column_mapping.get('new_columns', {})
        new_col_names = list(new_columns_mapping.values())
        if new_col_names and self.new_columns_combined is not None:
            new_cols_lf = self.new_columns_combined.lazy()
            source_lf = source_lf.with_row_index(name='_row_idx_')
            source_lf = source_lf.join(
                new_cols_lf.rename({'row_idx': '_row_idx_'}),
                on='_row_idx_',
                how='left'
            )
            source_lf = source_lf.drop('_row_idx_')
        
        # Append additions
        if self.additions is not None and len(self.additions) > 0:
            add_lf = self.additions.lazy()
            source_schema = source_lf.collect_schema()
            source_cols = source_schema.names()

            add_schema = add_lf.collect_schema()
            for col in source_cols:
                if col not in add_schema.names():
                    add_lf = add_lf.with_columns(pl.lit(None).alias(col))
                    add_schema = add_lf.collect_schema()

            add_lf = add_lf.select(source_cols)
            add_schema = add_lf.collect_schema()

            for col in source_cols:
                if col in add_schema.names():
                    add_type = add_schema[col]
                    src_type = source_schema[col]
                    if add_type != src_type and add_type != pl.Null:
                        add_lf = add_lf.with_columns(pl.col(col).cast(src_type))
                        add_schema = add_lf.collect_schema()
            source_lf = pl.concat([source_lf, add_lf])
        
        # Write to temporary Parquet using streaming sink (low memory)
        parquet_temp = output_path + ".parquet"
        source_lf.sink_parquet(parquet_temp)
        del source_lf
        
        # Now scan the Parquet file and get total row count (fast)
        lazy_df = pl.scan_parquet(parquet_temp)
        total_rows = lazy_df.select(pl.count()).collect().item()
        headers = lazy_df.collect_schema().names()
        
        self.progress_updated.emit(f"Writing {total_rows:,} rows to XLSX...")
        
        # Extract original header formats (read-only, fast, no column widths)
        header_formats = self._get_header_formats()
        
        # Write to xlsxwriter with constant_memory mode
        import xlsxwriter
        workbook = xlsxwriter.Workbook(output_path, {'constant_memory': True})
        worksheet = workbook.add_worksheet()
        
        # Format cache to avoid creating duplicate format objects
        fmt_cache = {}
        def get_fmt(col_idx):
            if col_idx not in header_formats:
                return None
            props = header_formats[col_idx]
            key = tuple(props.items())
            if key not in fmt_cache:
                fmt = workbook.add_format({
                    'font_name': props['font_name'],
                    'font_size': props['font_size'],
                    'bold': props['bold'],
                    'font_color': props['font_color'],
                    'bg_color': props['bg_color'],
                    'align': props['align'],
                    'valign': props['valign'],
                    'num_format': props['num_format'],
                    'border': 1,
                })
                fmt_cache[key] = fmt
            return fmt_cache[key]
        
        # Write headers with original formatting
        for col_idx, header in enumerate(headers):
            fmt = get_fmt(col_idx + 1)  # openpyxl columns are 1-indexed
            if fmt:
                worksheet.write(0, col_idx, header, fmt)
            else:
                worksheet.write(0, col_idx, header)
        
        # Write data rows in chunks
        chunk_size = 20000
        for start in range(0, total_rows, chunk_size):
            batch = lazy_df.slice(start, chunk_size).collect()
            self._write_batch_to_worksheet(worksheet, batch, start_row=start + 1)
            if start % (chunk_size * 10) == 0:
                self.progress_updated.emit(
                    f"Written {min(start + chunk_size, total_rows):,}/{total_rows:,} rows"
                )
            del batch
        
        workbook.close()
        
        # Clean up Parquet temporary file
        try:
            os.remove(parquet_temp)
        except OSError:
            pass
        
        self.progress_updated.emit("XLSX write complete.")
    
    def _write_batch_to_worksheet(self, worksheet, df, start_row):
        """Write a Polars DataFrame chunk to an xlsxwriter worksheet (optimized)."""
        for row_idx, row in enumerate(df.iter_rows()):
            worksheet.write_row(start_row + row_idx, 0, row)
    
    # ── CSV Writing ─────────────────────────────────────────────────────
    def _write_csv(self, output_path: str):
        """Write data to CSV using utf-8-sig encoding for Excel compatibility.
        
        Optimized: Uses Polars' native write_csv for maximum speed,
        then adds BOM via streaming copy to avoid memory peaks.
        Progress is reported as rows written (no total count).
        Temporary file is always cleaned up, even on errors.
        """
        self.progress_updated.emit("Building output data...")
        
        source_lf = self._get_lazy_source()
        source_lf = self._apply_updates(source_lf)
        
        # Add new columns (optimized: single join)
        new_columns_mapping = self.column_mapping.get('new_columns', {})
        new_col_names = list(new_columns_mapping.values())
        if new_col_names and self.new_columns_combined is not None:
            new_cols_lf = self.new_columns_combined.lazy()
            source_lf = source_lf.with_row_index(name='_row_idx_')
            source_lf = source_lf.join(
                new_cols_lf.rename({'row_idx': '_row_idx_'}),
                on='_row_idx_',
                how='left'
            )
            source_lf = source_lf.drop('_row_idx_')
        
        # Append additions
        if self.additions is not None and len(self.additions) > 0:
            add_lf = self.additions.lazy()
            source_schema = source_lf.collect_schema()
            source_cols = source_schema.names()

            add_schema = add_lf.collect_schema()
            for col in source_cols:
                if col not in add_schema.names():
                    add_lf = add_lf.with_columns(pl.lit(None).alias(col))
                    add_schema = add_lf.collect_schema()

            add_lf = add_lf.select(source_cols)
            add_schema = add_lf.collect_schema()

            for col in source_cols:
                if col in add_schema.names():
                    add_type = add_schema[col]
                    src_type = source_schema[col]
                    if add_type != src_type and add_type != pl.Null:
                        add_lf = add_lf.with_columns(pl.col(col).cast(src_type))
                        add_schema = add_lf.collect_schema()
            source_lf = pl.concat([source_lf, add_lf])
        
        # Collect all data at once (single lazy evaluation)
        self.progress_updated.emit("Collecting result...")
        result_df = source_lf.collect()
        self.progress_updated.emit(f"Writing {len(result_df):,} rows to CSV...")
        
        # Step 1: Write to a temporary file using fast Polars CSV writer
        temp_csv = output_path + ".tmp"
        try:
            result_df.write_csv(temp_csv, batch_size=20000)
            del result_df  # free memory
            
            # Step 2: Stream copy to final output with BOM prefix
            self.progress_updated.emit("Adding UTF-8 BOM...")
            with open(temp_csv, 'rb') as src, open(output_path, 'wb') as dst:
                dst.write(b'\xef\xbb\xbf')  # UTF-8 BOM
                shutil.copyfileobj(src, dst)
        finally:
            # Always clean up the temporary CSV file
            try:
                os.remove(temp_csv)
            except OSError:
                pass
        
        self.progress_updated.emit("CSV write complete.")
    
    def _get_lazy_source(self):
        """Obtain a LazyFrame from the target file, with memory-friendly scanning for CSV."""
        if self.target_path.endswith('.csv'):
            return pl.scan_csv(self.target_path, has_header=True, truncate_ragged_lines=True)
        else:
            # For Excel, we must eager read, then convert to lazy
            # This is a known limitation; large Excel files may consume memory here.
            df = pl.read_excel(self.target_path, engine='calamine')
            return df.lazy()
    
    @staticmethod
    def _detect_encoding(file_path):
        """Detect file encoding for CSV files."""
        import chardet
        with open(file_path, 'rb') as f:
            raw = f.read(50000)
            result = chardet.detect(raw)
            return result['encoding'] or 'utf-8'
    
    # ── Formula Restoration ─────────────────────────────────────────────
    def _restore_formulas(self, output_path: str):
        """Restore formulas detected in the original file."""
        formulas = self.formula_info.get('formulas', [])
        if not formulas:
            return
        
        self.progress_updated.emit("Restoring formulas...")
        
        wb = load_workbook(output_path)
        ws = wb.active
        
        for formula in formulas:
            row = formula['row']
            col_idx = formula['column_index']
            if row <= ws.max_row:
                cell = ws.cell(row=row, column=col_idx + 1)
                cell.value = formula['formula']
        
        wb.save(output_path)
        wb.close()
        
        self.progress_updated.emit("Formulas restored.")
